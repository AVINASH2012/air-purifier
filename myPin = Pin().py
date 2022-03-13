import openpyxl
import serial 
import tkinter as tk
from tkinter import messagebox 
import re

###########################################################################
Arduino_Serial = serial.Serial('com7',baudrate=115200,timeout=30)          
print("connected to: " + Arduino_Serial.portstr)                          
###########################################################################


class Command(object):
    # def gui(self):
    #     self.result1 = messagebox.askyesno(
    #             title = 'YesnoDemo',
    #             message = 'do you see output',
    #             detail = 'click no to quit'
    #         )
    #     return self.result1   


    def readcmd(self,value,i,cell_obj):
        # print(cell_obj.value)
        
        if i==2:
            Arduino_Serial.write(("factory exec 1\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            # try:
            while (1):
                arr= []
                data = (Arduino_Serial.readline())
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                finalValue = (finalValue).split("\n")
                for val in finalValue:
                    if re.search(r'^factory exec 1 0', val):
                        arr.append(val)
                        print(arr[0].replace('\x1b[0;32mHIFECLI> \x1b[0m ', ''))
                        temp = (arr[0].replace('\x1b[0;32mHIFECLI> \x1b[0m ', ''))
                        wb_obj2= sheet_obj.cell(row = i, column = 3)
                        wb_obj2.value = temp
                        break 
            
                if arr != []:
                    break     
                    # print ([ val for val in finalValue if re.search(r'^factory', val) ])
             
            # self.gui() 
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
                
            )
           
            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "success"
                print(wb_obj3.value)

            if temp and self.result1:
                print('pass')  
                wb_obj3= sheet_obj.cell(row = i, column = 5)
                wb_obj3.value = 'pass'  
                
        if cell_obj.value == 'factory exec 2':
            Arduino_Serial.write(("factory exec 2\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
               
                data = (Arduino_Serial.readline())
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue)
               
            except:#requests.exceptions.ConnectTimeout:
                print('Failed')

            # self.gui()
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
            )    
            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 2"
                #print(wb_obj3.value)
                print('success')
            else:
                exit()       
              
            

        if cell_obj.value == 'factory exec 3':
            Arduino_Serial.write(("factory exec 3\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
                data = (Arduino_Serial.readline())
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue) 
              
            except: #requests.exceptions.ConnectTimeout:
                print('Failed')

            # self.gui()
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
            )  

            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 3"
                #print(wb_obj3.value)
                print('success')
            else:
                exit()       
            

        if cell_obj.value == 'factory exec 4':
            Arduino_Serial.write(("factory exec 4\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
                data = (Arduino_Serial.readline()) 
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue)
            except: #requests.exceptions.ConnectTimeout:
                print('Failed')

            # self.gui()
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
            ) 

            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 4"
                #print(wb_obj3.value)
                print('success')
            else:
                exit()       

        if cell_obj.value == 'factory exec 5':
            Arduino_Serial.write(("factory exec 5\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
                data = (Arduino_Serial.readline()) 
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue)
            except: #requests.exceptions.ConnectTimeout:
                print('Failed')
            
            # self.gui()
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = 'do you see shut down button LED',
                detail = 'click no to quit'
            )

            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 5"
                #print(wb_obj3.value)
                print('success')
            else:
                exit()
        if cell_obj.value == 'factory exec 6':
            Arduino_Serial.write(("factory exec 6\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
                data = (Arduino_Serial.readline()) 
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue)
            except: #requests.exceptions.ConnectTimeout:
                print('Failed')
            
            # self.gui()'
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
            )

            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 6"
                #print(wb_obj3.value)
                print('success')


        if cell_obj.value == 'factory exec 7':
            Arduino_Serial.write(("factory exec 7\r\n").encode())
            cell_obj2 = sheet_obj.cell(row= i, column =2)
            print(cell_obj2.value)
            try:
                data = (Arduino_Serial.readline()) 
                finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
                print (finalValue)
            except: #requests.exceptions.ConnectTimeout:
                print('Failed')
            
            # self.gui()
            temp = sheet_obj.cell(row= i, column =6)
            self.result1 = messagebox.askyesno(
                title = 'Havells',
                message = temp.value,
                detail = 'click no to quit'
            )

            if self.result1 :
                wb_obj3= sheet_obj.cell(row = i, column = 4)
                wb_obj3.value = "factory exec 7"
                #print(wb_obj3.value)
                print('success')        
            else:
                exit() 



        else:
            wb_obj.save("C:\\Users\\adity\\hello2.xlsx")
           

print('---------------command starts--------------------')

###################################
val = str(input())
###################################
if val == "factory start":
    Arduino_Serial.write(("factory start\r\n").encode())
    data = (Arduino_Serial.readline())
    finalValue = data.decode('utf-7').replace('\r','').replace('\n','')
    print(finalValue)
###################################   
    mycomm = Command()
    wb_obj =  openpyxl.load_workbook("C:\\Users\\adity\\hello2.xlsx")
    sheet_obj= wb_obj.active
    max_row = sheet_obj.max_row
    i = 1
    # while (i<max_row):
    #     val1 = str(input())
    #     for i in range(1,max_row):
    #         cell_obj = sheet_obj.cell(row = i,column=1)
    #         if str(val1) == str(cell_obj.value):
    #             value = (cell_obj.value)
    #             mycomm.readcmd(value,i,cell_obj)
    #     if str(val1) == 'factory stop' :
    #         Arduino_Serial.write(("factory stop\r\n").encode())
    #         break    
    while (i<max_row):
        for i in range(2,max_row):
            cell_obj = sheet_obj.cell(row = i,column=1)
            value = (cell_obj.value)
            print(value)
            mycomm.readcmd(value,i,cell_obj)
            i+=1
            if i== max_row:
                i=1

else:
    print('Failed')

          